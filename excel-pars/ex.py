from openpyxl import load_workbook

from AllSimpleDB import MySql


db = MySql('data_excel', 'root', '28108313art15', 'localhost')
db.executeQuery("DELETE FROM excel_details")

book = load_workbook(filename='excel-pars2/main.xlsx', data_only='True')
sheet = book['Нержа']

list_str_db = ""
index = 1

for row in range(1, sheet.max_row + 1): #адрес ячейки не может быть 0, поэтому от 1
    errors_count = 0
    name = str(sheet[row][0].value)

    try:   #пытаемся конвертировать во float
        width = str(float(str(sheet[row][1].value)))
    except: #иначе строка
        width = str(sheet[row][1].value)
        errors_count += 1

    try:
        length = str(float(str(sheet[row][2].value)))[0:4]
    except:
        length = str(sheet[row][2].value)
        errors_count += 1

    try:
        amount = str(int(str(sheet[row][3].value)))[0:4]
    except:
        amount = str(sheet[row][3].value)
        errors_count += 1

    try:
        square_meter = str(float(str(sheet[row][4].value)))[0:4]
    except:
        square_meter = str(sheet[row][4].value)
        errors_count += 1

    try:
        summa = str(float(str(sheet[row][5].value)))[0:4]
    except:
        summa = str(sheet[row][5].value)
        errors_count += 1

    # if width == 'None' and length == 'None' and amount == 'None' and square_meter == 'None' and summa == 'None' or errors_count == 5:
    #     continue
    if width == 'None' and length == 'None' and amount == 'None' and square_meter == 'None'  or errors_count == 5:
        continue

    # list_str_db += f"{index, name, width if width != 'None' else 0, length if length != 'None' else 0, amount if amount != 'None' else 0, square_meter if square_meter != 'None' else 0, summa if summa != 'None' else 0},"
    # index += 1
    list_str_db += f"{index, name, width if width != 'None' else 0, length if length != 'None' else 0, amount if amount != 'None' else 0, square_meter if square_meter != 'None' else 0},"
    index += 1

    print(index, name, width, length, square_meter)

list_str_db = list_str_db[:-1] #убрать последнюю запятую в конце

db.executeQuery(f"""
    INSERT INTO
        excel_details (id, name, width, length, amount, square_meter)
    VALUE
        {list_str_db}
    """) 
